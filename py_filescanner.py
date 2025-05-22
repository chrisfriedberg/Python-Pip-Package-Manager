#!/usr/bin/env python
"""
Python File Scanner - A tool to scan for Python files and manage them.
Refactored for improved error handling, clarity, and queued deletions.
"""
import os
import sys
import datetime
import logging
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication, QWidget, QFileDialog, QPushButton, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox, QDialog, QHeaderView,
    QAbstractItemView, QLabel, QTextEdit, QStyle
)
from PySide6.QtCore import Qt, QDate # Added QDate
import subprocess
import send2trash

# Set up logging to a file and to the debug window
# File logging
log_file_path = Path(__file__).parent / "file_scanner.log"
file_handler = logging.FileHandler(log_file_path)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(module)s - %(message)s'))

# Console/Stream logging (for terminal, if run from there)
stream_handler = logging.StreamHandler()
stream_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logging.basicConfig(level=logging.DEBUG, handlers=[file_handler, stream_handler])
logger = logging.getLogger(__name__)


class DebugWindow(QDialog):
    """A debug window to display log messages from the application."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Debug Log")
        self.setMinimumSize(800, 400) 
        self.setObjectName("DebugWindow") 

        layout = QVBoxLayout(self) 
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setObjectName("LogText")
        layout.addWidget(self.log_text)

        self.clear_button = QPushButton("Clear Log")
        self.clear_button.clicked.connect(self.clear_log_display)
        layout.addWidget(self.clear_button)

        self.setLayout(layout)
        logger.info("DebugWindow initialized.")

    def log_message(self, message, level="INFO"):
        """Add a message to the log display and the main logger."""
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        formatted_message = f"{timestamp} - {level} - {message}"
        self.log_text.append(formatted_message)
        
        if level.upper() == "DEBUG": logger.debug(message)
        elif level.upper() == "INFO": logger.info(message)
        elif level.upper() == "WARNING": logger.warning(message)
        elif level.upper() == "ERROR": logger.error(message)
        elif level.upper() == "CRITICAL": logger.critical(message)
        else: logger.info(message)

    def clear_log_display(self):
        """Clear the log display in the debug window."""
        self.log_text.clear()
        self.log_message("Debug log display cleared.", "INFO")


class FileScanner:
    """Handles the file scanning logic separate from the UI."""

    def __init__(self, debug_logger_func=None):
        self.scanned_files = []
        self.debug_log = debug_logger_func if debug_logger_func else logger.info
        self.debug_log("FileScanner initialized.", "INFO")

    def _walk_error_handler(self, os_error):
        error_message = f"Access error during scan: {os_error.filename} - {os_error.strerror}. Skipping this path."
        self.debug_log(error_message, "WARNING")

    def scan_directory(self, directory):
        self.debug_log(f"Starting scan of directory: {directory}", "INFO")
        self.scanned_files = []
        files_processed_count = 0
        python_files_found_count = 0

        if not Path(directory).is_dir():
            self.debug_log(f"Provided path is not a valid directory: {directory}", "ERROR")
            return []

        try:
            for root, dirs, files in os.walk(directory, onerror=self._walk_error_handler, followlinks=False):
                self.debug_log(f"Scanning: {root} (Found {len(dirs)} subdirs, {len(files)} files in current dir)", "DEBUG")
                for file_name in files:
                    files_processed_count += 1
                    if file_name.lower().endswith((".py", ".pyw")): # Only scan for .py and .pyw files
                        python_files_found_count += 1
                        try:
                            full_path = Path(root) / file_name
                            stat_info = full_path.stat()
                            created_time = datetime.datetime.fromtimestamp(stat_info.st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                            modified_time = datetime.datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                            file_info = {
                                "directory": str(full_path.parent),
                                "filename": full_path.name,
                                "created": created_time,
                                "modified": modified_time,
                                "status": "File Found" # Initial status
                            }
                            self.scanned_files.append(file_info)
                            if python_files_found_count <= 10:
                                self.debug_log(f"Found Python file: {full_path.name} in {full_path.parent}", "DEBUG")
                        except FileNotFoundError:
                            self.debug_log(f"File not found during stat: {full_path}. Potentially a broken symlink or race condition.", "WARNING")
                        except PermissionError:
                            self.debug_log(f"Permission error accessing metadata for: {full_path}", "WARNING")
                        except Exception as e:
                            self.debug_log(f"Error processing file metadata for {file_name} in {root}: {e}", "ERROR")
        except Exception as e:
            self.debug_log(f"General error during directory scan operation: {e}", "CRITICAL")

        self.debug_log(f"Scan complete. Processed {files_processed_count} total items. Found {len(self.scanned_files)} Python files.", "INFO")
        return self.scanned_files

    def delete_file_to_trash(self, file_path_str):
        file_path = Path(file_path_str)
        if not file_path.exists():
            self.debug_log(f"Attempted to delete non-existent file: {file_path_str}", "ERROR")
            return False
        try:
            send2trash.send2trash(file_path_str)
            self.debug_log(f"Successfully sent to trash: {file_path_str}", "INFO")
            return True
        except send2trash.TrashPermissionError as e:
            self.debug_log(f"Permission error sending to trash {file_path_str}: {e}", "ERROR")
        except Exception as e:
            self.debug_log(f"Error sending file to trash {file_path_str}: {e}", "ERROR")
        return False

    def try_delete_empty_dir_to_trash(self, directory_str, all_py_files_not_in_current_deletion_batch):
        """
        Try to delete a directory by sending it to trash, subject to conditions:
        1. Not a critical OS/User directory.
        2. No other application-tracked .py/.pyw files (from all_py_files_not_in_current_deletion_batch) 
           remain within this directory path.
        3. The directory is physically empty on disk.
        """
        directory = Path(directory_str)
        self.debug_log(f"Attempting conditional delete for directory: {directory}", "DEBUG")

        home_dir = Path.home()
        # Expand the list of user-specific critical subfolders
        user_critical_subfolders = [
            "Documents", "Desktop", "Downloads", "Pictures", "Music", "Videos",
            "AppData", ".config", ".local", # Common hidden config folders
            "OneDrive", "Dropbox", "Google Drive" # Common cloud sync folders
        ]
        
        critical_paths = [Path(p) for p in [
            os.getenv("SystemRoot", "C:\\Windows"),
            os.getenv("ProgramFiles", "C:\\Program Files"),
            os.getenv("ProgramFiles(x86)", "C:\\Program Files (x86)"),
            home_dir 
        ] if p] # Ensure p is not None
        
        for folder_name in user_critical_subfolders:
            critical_paths.append(home_dir / folder_name)
        
        try:
            normalized_directory_to_check = directory.resolve()
        except Exception: 
            normalized_directory_to_check = directory

        is_critical_dir_or_child = any(
            crit_path.exists() and (normalized_directory_to_check == crit_path.resolve() or crit_path.resolve() in normalized_directory_to_check.parents)
            for crit_path in critical_paths
        )
        if is_critical_dir_or_child:
            self.debug_log(f"Skipping deletion attempt for critical OS/User directory or its child: {directory_str}", "WARNING")
            return

        remaining_py_files_in_dir_path = 0
        normalized_target_dir_path = normalized_directory_to_check

        # Check against files NOT in the current deletion batch
        for file_info in all_py_files_not_in_current_deletion_batch: 
            try:
                # Ensure 'directory' and 'filename' keys exist
                if "directory" not in file_info or "filename" not in file_info:
                    self.debug_log(f"Skipping file_info with missing keys: {file_info}", "WARNING")
                    continue

                file_parent_dir = Path(file_info["directory"]).resolve()
                if file_parent_dir == normalized_target_dir_path or normalized_target_dir_path in file_parent_dir.parents:
                    if file_info["filename"].lower().endswith((".py", ".pyw")):
                        remaining_py_files_in_dir_path += 1
                        self.debug_log(f"Found remaining PY file: {file_info['filename']} in {file_parent_dir}", "DEBUG")
                        break 
            except Exception as e:
                self.debug_log(f"Error processing file_info during remaining PY check for {directory_str}: {file_info} - {e}", "ERROR")

        if remaining_py_files_in_dir_path > 0:
            self.debug_log(f"Directory {directory_str} not deleted: {remaining_py_files_in_dir_path} other Python files still tracked by the app (and not in current deletion batch) within this path.", "INFO")
            return

        self.debug_log(f"No other app-tracked Python files (outside current batch) found in {directory_str} path. Proceeding to physical emptiness check.", "INFO")

        try:
            if directory.is_dir(): 
                # Check for physical emptiness (no files of any type, no non-empty subdirs)
                is_physically_empty = True
                for item in directory.iterdir():
                    if item.is_file(): # Any file makes it not empty
                        is_physically_empty = False
                        break
                    if item.is_dir(): # A subdirectory
                        try:
                            if any(item.iterdir()): # If subdir is not empty
                                is_physically_empty = False
                                break
                        except PermissionError: # Cannot access subdir to check if empty
                            is_physically_empty = False # Assume not empty for safety
                            self.debug_log(f"Permission error checking emptiness of subdir {item} in {directory_str}. Assuming not empty.", "WARNING")
                            break
                
                if is_physically_empty:
                    send2trash.send2trash(directory_str)
                    self.debug_log(f"Successfully sent empty directory to trash: {directory_str}", "INFO")
                else:
                    self.debug_log(f"Directory {directory_str} not deleted: It was the last app-tracked Python project folder in this path (considering current batch), but it still physically contains other files or non-empty subdirectories.", "INFO")
        except send2trash.TrashPermissionError as e:
            self.debug_log(f"Permission error sending directory to trash {directory_str}: {e}", "ERROR")
        except FileNotFoundError:
             self.debug_log(f"Directory {directory_str} not found during physical check. No action taken.", "INFO")
        except Exception as e:
            self.debug_log(f"Error during final directory deletion attempt for {directory_str}: {e}", "WARNING")


class ReviewDialog(QDialog):
    """Dialog for reviewing files before adding them to the deletion queue."""

    def __init__(self, items_to_delete, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Review & Confirm Files for Deletion Queue")
        self.setMinimumSize(1000, 600)  # Make dialog larger to match main window better
        self.items_to_delete_initially = list(items_to_delete) 
        self._debug_log_func = parent.debug_window.log_message if hasattr(parent, 'debug_window') and hasattr(parent.debug_window, 'log_message') else logger.info
        self._debug_log_func(f"REVIEW DEBUG: ReviewDialog initialized with {len(self.items_to_delete_initially)} items to review", "INFO")
        
        # Debug log each item for visibility
        for i, (dir_path, file_name) in enumerate(self.items_to_delete_initially):
            self._debug_log_func(f"REVIEW DEBUG: Item {i+1}: {dir_path} / {file_name}", "INFO")
            
        self._setup_ui()
        self._populate_table()
        self._debug_log_func(f"ReviewDialog initialized with {len(self.items_to_delete_initially)} items.", "INFO")

    def _setup_ui(self):
        # Main layout
        self.layout = QVBoxLayout(self)
        
        # Instructions label
        instructions = QLabel("<b>Review files to be added to the deletion queue.</b> Uncheck items you do NOT want to queue.")
        instructions.setTextFormat(Qt.RichText)
        self.layout.addWidget(instructions)

        # Create the table widget - mirroring the main app's table setup
        self.table = QTableWidget(0, 3) 
        self.table.setHorizontalHeaderLabels(["Queue?", "Directory", "Filename"])
        
        # Table selection and behavior settings - similar to main app
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)  # Alternating row colors like main app
        
        # Column resize modes like the main app
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        
        # Set initial column widths for better visibility
        self.table.setColumnWidth(0, 80)    # Queue? column
        self.table.setColumnWidth(1, 500)   # Directory column (wider)
        self.table.setColumnWidth(2, 300)   # Filename column
        
        self.layout.addWidget(self.table)

        # Button layout
        dialog_button_layout = QHBoxLayout()
        
        # Export button
        self.export_button = QPushButton("Export List for Deletion")
        self.export_button.clicked.connect(self.export_deletion_list)
        dialog_button_layout.addWidget(self.export_button)
        dialog_button_layout.addStretch()

        # Confirm button
        self.confirm_button = QPushButton("Add Checked to Deletion Queue")
        self.confirm_button.clicked.connect(self.accept)
        self.confirm_button.setDefault(True) 
        dialog_button_layout.addWidget(self.confirm_button)

        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        dialog_button_layout.addWidget(self.cancel_button)
        
        self.layout.addLayout(dialog_button_layout)
        self.setLayout(self.layout)

    def _populate_table(self):
        """Populate table with the items to be reviewed, mirror of main app's populate_table_view"""
        # Clear table and set row count
        self.table.setRowCount(0)  # Clear table first
        total_items = len(self.items_to_delete_initially)
        self._debug_log_func(f"REVIEW DEBUG: About to populate table with {total_items} items", "INFO")
        
        # Disable sorting during population
        self.table.setSortingEnabled(False)
        
        # Set exact row count
        self.table.setRowCount(total_items)
        
        # Populate each row
        for row_idx, (dir_path, file_name) in enumerate(self.items_to_delete_initially):
            self._debug_log_func(f"REVIEW DEBUG: Setting row {row_idx} with {dir_path} / {file_name}", "INFO")
            
            # Checkbox item for Queue? column
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox_item.setCheckState(Qt.Checked) 
            
            # Directory item
            dir_item = QTableWidgetItem(dir_path)
            dir_item.setFlags(dir_item.flags() & ~Qt.ItemIsEditable)
            
            # Filename item
            file_item = QTableWidgetItem(file_name)
            file_item.setFlags(file_item.flags() & ~Qt.ItemIsEditable)
            
            # Set each item in the table
            self.table.setItem(row_idx, 0, checkbox_item)
            self.table.setItem(row_idx, 1, dir_item)
            self.table.setItem(row_idx, 2, file_item)
            
            # Verify items were set correctly
            verify_checkbox = self.table.item(row_idx, 0)
            verify_dir = self.table.item(row_idx, 1)
            verify_file = self.table.item(row_idx, 2)
            
            if verify_checkbox and verify_dir and verify_file:
                self._debug_log_func(f"REVIEW DEBUG: Successfully set row {row_idx}", "INFO")
            else:
                self._debug_log_func(f"REVIEW DEBUG: Failed to set items for row {row_idx}", "ERROR")
        
        # Re-enable sorting after population
        self.table.setSortingEnabled(True)
        
        # Final check of row count
        final_rows = self.table.rowCount()
        self._debug_log_func(f"REVIEW DEBUG: Table populated, final row count: {final_rows}", "INFO")
        
        # Make column width adjustments
        self.table.setColumnWidth(0, 80)        # Queue? column
        self.table.setColumnWidth(1, 500)       # Directory column
        self.table.setColumnWidth(2, 300)       # Filename column

    def get_items_confirmed_for_queue(self):
        """Get items that are still checked to be added to the deletion queue."""
        confirmed_items = []
        total_rows = self.table.rowCount()
        self._debug_log_func(f"REVIEW DEBUG: Retrieving confirmed items from {total_rows} rows", "INFO")
        
        for row in range(total_rows):
            self._debug_log_func(f"REVIEW DEBUG: Processing row {row} for confirmation", "INFO")
            
            # Check that all necessary items exist
            checkbox_item = self.table.item(row, 0)
            dir_item = self.table.item(row, 1)
            file_item = self.table.item(row, 2)
            
            # Log the state of each item for debugging
            has_checkbox = checkbox_item is not None
            has_dir = dir_item is not None
            has_file = file_item is not None
            self._debug_log_func(f"REVIEW DEBUG: Row {row} items - Checkbox: {has_checkbox}, Dir: {has_dir}, File: {has_file}", "INFO")
            
            # Only process if all items exist
            if checkbox_item and dir_item and file_item:
                # Check if the checkbox is checked
                is_checked = checkbox_item.checkState() == Qt.Checked
                self._debug_log_func(f"REVIEW DEBUG: Row {row} checkbox state: {'Checked' if is_checked else 'Unchecked'}", "INFO")
                
                if is_checked:
                    dir_path = dir_item.text()
                    file_name = file_item.text()
                    self._debug_log_func(f"REVIEW DEBUG: Adding row {row} to confirmed items: {dir_path} / {file_name}", "INFO")
                    confirmed_items.append((dir_path, file_name))
                else:
                    self._debug_log_func(f"REVIEW DEBUG: Row {row} is unchecked - skipping", "INFO")
            else:
                # Access the items from the original list for this row
                if row < len(self.items_to_delete_initially):
                    orig_dir, orig_file = self.items_to_delete_initially[row]
                    self._debug_log_func(f"REVIEW DEBUG: Row {row} missing table items, but found in original list: {orig_dir} / {orig_file}", "WARNING")
                    confirmed_items.append((orig_dir, orig_file))
                else:
                    self._debug_log_func(f"REVIEW DEBUG: Row {row} missing data and not found in original list", "WARNING")
                
        self._debug_log_func(f"REVIEW DEBUG: ReviewDialog returning {len(confirmed_items)} items confirmed for deletion queue.", "INFO")
        return confirmed_items

    def export_deletion_list(self):
        items_to_export = self.get_items_confirmed_for_queue()
        if not items_to_export:
            QMessageBox.information(self, "No Items to Export", "No items are currently checked to be queued for deletion.")
            return

        default_filename = f"DeletionLog_{QDate.currentDate().toString('yyyy-MM-dd')}.txt"
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Deletion Log", 
            str(Path.home() / default_filename), # Start in user's home directory
            "Text Files (*.txt);;All Files (*)"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(f"Deletion Log - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 40 + "\n")
                    f.write("Files to be queued for deletion:\n\n")
                    for dir_path, file_name in items_to_export:
                        f.write(f"Directory: {dir_path}\nFilename:  {file_name}\n---\n")
                QMessageBox.information(self, "Export Successful", f"Deletion list exported to:\n{file_path}")
                self._debug_log_func(f"Deletion list exported to {file_path}", "INFO")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", f"Could not export deletion list: {e}")
                self._debug_log_func(f"Error exporting deletion list: {e}", "ERROR")


class FileScannerApp(QWidget):
    """Main application class for the Python File Scanner."""

    def __init__(self):
        super().__init__()
        self.scanner = None 
        self.current_scanned_files_data = [] 
        self.files_queued_for_deletion = set()
        self.is_exit_after_deletion = False  # Flag to track if we're exiting after deletion
        
        self.debug_window = DebugWindow(self) 
        self.scanner = FileScanner(debug_logger_func=self.debug_window.log_message)
        
        self.debug_window.log_message("Application FileScannerApp initialized.", "INFO")
        self.setWindowTitle("Python File Scanner")
        self.setMinimumSize(1100, 600) # Adjusted for new buttons
        self.setup_ui()
        self.update_queue_counter() # Initialize counter display
        self.debug_window.log_message("Main UI setup complete.", "INFO")

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        top_controls_layout = QHBoxLayout()
        scan_button = QPushButton("Select Directory & Scan")
        scan_button.setIcon(self.style().standardIcon(QStyle.SP_DirOpenIcon))
        scan_button.clicked.connect(self.select_and_scan_directory)
        top_controls_layout.addWidget(scan_button)

        self.debug_button = QPushButton("Toggle Debug Window")
        self.debug_button.setCheckable(True)
        self.debug_button.clicked.connect(self.toggle_debug_window)
        top_controls_layout.addWidget(self.debug_button)
        if self.debug_window.isVisible(): self.debug_button.setChecked(True)
        top_controls_layout.addStretch(1)
        main_layout.addLayout(top_controls_layout)

        search_header_label = QLabel("<b>Filter Results:</b>")
        search_header_label.setTextFormat(Qt.RichText)
        main_layout.addWidget(search_header_label)
        search_layout = QHBoxLayout()
        self.search_boxes = {} 
        column_labels = ["Directory", "Filename", "Created", "Modified", "Status"]
        for idx, label_text in enumerate(column_labels):
            search_box = QLineEdit()
            search_box.setPlaceholderText(f"Filter by {label_text}...")
            search_box.textChanged.connect(self.filter_table_view)
            self.search_boxes[label_text.lower()] = search_box 
            search_layout.addWidget(search_box)
        main_layout.addLayout(search_layout)

        self.table = QTableWidget(0, len(column_labels))
        self.table.setHorizontalHeaderLabels(column_labels)
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers) 
        self.table.cellDoubleClicked.connect(self.open_file_directory_action)
        
        # Make columns user-adjustable
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        # Set initial column widths
        header.setDefaultSectionSize(120)        # Default size
        header.resizeSection(0, 350)             # Directory
        header.resizeSection(1, 200)             # Filename
        header.resizeSection(2, 150)             # Created
        header.resizeSection(3, 150)             # Modified
        header.resizeSection(4, 100)             # Status
        
        self.table.setAlternatingRowColors(True) 
        main_layout.addWidget(self.table)

        # --- Action Buttons Layout ---
        action_buttons_layout = QHBoxLayout()
        
        # Left side buttons
        self.queue_delete_button = QPushButton("Queue Selected for Deletion...")
        self.queue_delete_button.setIcon(self.style().standardIcon(QStyle.SP_DialogApplyButton)) 
        self.queue_delete_button.clicked.connect(self.queue_selected_files_for_deletion)
        action_buttons_layout.addWidget(self.queue_delete_button)
        
        self.clear_queued_button = QPushButton("Clear Queued Deletions")
        self.clear_queued_button.setIcon(self.style().standardIcon(QStyle.SP_DialogCancelButton))
        self.clear_queued_button.clicked.connect(self.clear_all_queued_deletions)
        action_buttons_layout.addWidget(self.clear_queued_button)
        
        # Queue counter and Execute Button in a horizontal layout
        queue_execute_layout = QHBoxLayout()
        self.queue_counter_label = QLabel("0")
        self.queue_counter_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
        self.queue_counter_label.setAlignment(Qt.AlignCenter)
        self.queue_counter_label.setMinimumWidth(30)
        
        self.execute_delete_button = QPushButton("Execute Deletions & Exit")
        self.execute_delete_button.setIcon(self.style().standardIcon(QStyle.SP_TrashIcon))
        self.execute_delete_button.setStyleSheet("background-color: #ffcccc; color: #a00000; font-weight: bold;")
        self.execute_delete_button.clicked.connect(self.execute_queued_deletions_and_exit)
        
        queue_execute_layout.addWidget(self.execute_delete_button)
        queue_execute_layout.addWidget(self.queue_counter_label)
        queue_execute_layout.setSpacing(5)
        
        action_buttons_layout.addLayout(queue_execute_layout)
        
        # Middle stretch
        action_buttons_layout.addStretch(1)
        
        # Right side buttons
        self.export_button = QPushButton("Export Current List")
        self.export_button.setIcon(self.style().standardIcon(QStyle.SP_FileDialogContentsView))
        self.export_button.clicked.connect(self.export_current_list)
        action_buttons_layout.addWidget(self.export_button)
        
        self.close_app_button = QPushButton("Close Application") 
        self.close_app_button.setIcon(self.style().standardIcon(QStyle.SP_DialogCloseButton))
        self.close_app_button.clicked.connect(self.close_application_handler)
        action_buttons_layout.addWidget(self.close_app_button)
        main_layout.addLayout(action_buttons_layout)

        self.setLayout(main_layout)

    def toggle_debug_window(self, checked):
        if checked:
            self.debug_window.show()
            self.debug_window.raise_() 
            self.debug_window.activateWindow() 
            self.debug_window.log_message("Debug window shown.", "INFO")
        else:
            self.debug_window.hide()
            self.debug_window.log_message("Debug window hidden.", "INFO")

    def select_and_scan_directory(self):
        if self.files_queued_for_deletion:
            reply = QMessageBox.question(self, "Unsaved Changes",
                                         "You have files queued for deletion. Scanning a new directory will clear this queue. Continue?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return

        self.debug_window.log_message("Select directory button clicked.", "INFO")
        start_dir = str(Path.home())
        directory = QFileDialog.getExistingDirectory(self, "Select Directory to Scan", start_dir)
        if not directory:
            self.debug_window.log_message("Directory selection cancelled by user.", "INFO")
            return
        self.debug_window.log_message(f"Directory selected for scan: {directory}", "INFO")
        self.table.setRowCount(0) 
        self.table.setCurrentItem(None) 
        self.current_scanned_files_data = [] 
        self.files_queued_for_deletion.clear() 
        self.update_queue_counter() # Update counter after clearing queue

        QApplication.setOverrideCursor(Qt.WaitCursor)
        self.debug_window.log_message("Starting file scan operation...", "INFO")
        found_files_data = self.scanner.scan_directory(directory)
        # Store the original data. The 'status' in this list will be updated if items are queued.
        self.current_scanned_files_data = [dict(item) for item in found_files_data] # Make copies
        
        self.debug_window.log_message(f"Scan finished. Populating table with {len(found_files_data)} files.", "INFO")
        self.populate_table_view() # Will use self.current_scanned_files_data
        QApplication.restoreOverrideCursor()
        self.debug_window.log_message("Table population complete. Cursor restored.", "INFO")
        if not found_files_data:
            QMessageBox.information(self, "Scan Complete", "No Python files (.py, .pyw) found in the selected directory or its subdirectories.")
        else:
            QMessageBox.information(self, "Scan Complete", f"Found {len(found_files_data)} Python files.")

    def populate_table_view(self): # No longer takes files_data_list as arg
        """Populates the main table using self.current_scanned_files_data."""
        self.table.setRowCount(0) 
        if not self.current_scanned_files_data:
            self.debug_window.log_message("No files data to populate in table.", "WARNING")
            self.table.setCurrentItem(None) 
            return

        self.table.setSortingEnabled(False) 
        self.table.setRowCount(len(self.current_scanned_files_data))
        for row_idx, file_entry in enumerate(self.current_scanned_files_data):
            try:
                dir_val = file_entry.get("directory", "N/A")
                file_val = file_entry.get("filename", "N/A")
                created_val = file_entry.get("created", "N/A")
                modified_val = file_entry.get("modified", "N/A")
                
                # Determine status based on whether it's in the deletion queue
                current_status = "File Found"
                if (dir_val, file_val) in self.files_queued_for_deletion:
                    current_status = "Queued for Deletion"
                else: # Ensure the status in the data reflects "File Found" if not queued
                    file_entry["status"] = "File Found"


                self.table.setItem(row_idx, 0, QTableWidgetItem(dir_val))
                self.table.setItem(row_idx, 1, QTableWidgetItem(file_val))
                self.table.setItem(row_idx, 2, QTableWidgetItem(created_val))
                self.table.setItem(row_idx, 3, QTableWidgetItem(modified_val))
                
                status_item = QTableWidgetItem(current_status)
                if current_status == "File Found": status_item.setForeground(Qt.darkGreen) 
                elif current_status == "Queued for Deletion": status_item.setForeground(Qt.red)
                self.table.setItem(row_idx, 4, status_item)
            except Exception as e:
                self.debug_window.log_message(f"Error populating table row {row_idx} with data {file_entry}: {e}", "ERROR")
        
        self.table.setSortingEnabled(True) 
        self.table.setCurrentItem(None) 
        self.debug_window.log_message(f"Table populated/refreshed with {self.table.rowCount()} rows.", "DEBUG")

    def filter_table_view(self):
        # This method now implicitly uses self.current_scanned_files_data for filtering
        # and then calls populate_table_view to show filtered results,
        # or directly manipulates row visibility based on the full self.current_scanned_files_data.
        # For simplicity with status, we'll repopulate based on a filtered list.

        search_terms = {key: box.text().lower() for key, box in self.search_boxes.items()}
        active_search_terms = {k: v for k, v in search_terms.items() if v}

        if not active_search_terms: # If no filters, show all
            self.populate_table_view() # Uses the full current_scanned_files_data
            return

        filtered_data = []
        column_map_keys = {"directory": "directory", "filename": "filename", "created": "created", "modified": "modified", "status": "status"}

        for file_entry in self.current_scanned_files_data:
            match = True
            for search_key_lower, term in active_search_terms.items():
                data_key = column_map_keys.get(search_key_lower)
                if not data_key: continue

                # Special handling for status if it's derived (like from files_queued_for_deletion)
                value_to_check = ""
                if data_key == "status":
                    value_to_check = "Queued for Deletion" if (file_entry.get("directory"), file_entry.get("filename")) in self.files_queued_for_deletion else "File Found"
                else:
                    value_to_check = file_entry.get(data_key, "").lower()
                
                if term not in value_to_check:
                    match = False
                    break
            if match:
                filtered_data.append(file_entry)
        
        # Temporarily populate with filtered data, but don't overwrite self.current_scanned_files_data
        # This approach for filtering by repopulating is a bit heavy.
        # A better way is to hide/show rows from the full table.
        # Reverting to row hiding for filter_table_view:
        
        self.table.setSortingEnabled(False) # Performance
        for row_idx in range(self.table.rowCount()):
            should_hide_row = False
            # Get data directly from table items for filtering visibility
            # This assumes table is already populated with full self.current_scanned_files_data
            # If populate_table_view is called after every filter text change, then this is fine.
            # Let's ensure populate_table_view is called to reflect the correct data first.
            # No, this is wrong. filter_table_view should HIDE rows, not repopulate.

            # Corrected filter logic:
            # Iterate through self.current_scanned_files_data, find corresponding row, and hide/show
            # This is complex if table is sorted. Simpler: iterate table rows.
            
            for search_key_lower, term in active_search_terms.items():
                col_idx = -1
                if search_key_lower == "directory": col_idx = 0
                elif search_key_lower == "filename": col_idx = 1
                elif search_key_lower == "created": col_idx = 2
                elif search_key_lower == "modified": col_idx = 3
                elif search_key_lower == "status": col_idx = 4
                
                if col_idx == -1: continue

                item = self.table.item(row_idx, col_idx)
                if not item or term not in item.text().lower():
                    should_hide_row = True
                    break
            self.table.setRowHidden(row_idx, should_hide_row)
        self.table.setSortingEnabled(True)
        self.table.setCurrentItem(None) 
        # Log count of visible rows
        visible_count = sum(1 for i in range(self.table.rowCount()) if not self.table.isRowHidden(i))
        self.debug_window.log_message(f"Table filtered. Visible rows: {visible_count}", "DEBUG")


    def open_file_directory_action(self, row, column):
        try:
            dir_item = self.table.item(row, 0) 
            if dir_item:
                dir_path_str = dir_item.text()
                dir_path = Path(dir_path_str)
                if dir_path.is_dir():
                    self.debug_window.log_message(f"Opening directory: {dir_path_str}", "INFO")
                    if sys.platform == "win32": os.startfile(dir_path_str)
                    elif sys.platform == "darwin": subprocess.Popen(["open", dir_path_str])
                    else: subprocess.Popen(["xdg-open", dir_path_str])
                else:
                    self.debug_window.log_message(f"Path is not a valid directory: {dir_path_str}", "WARNING")
                    QMessageBox.warning(self, "Invalid Path", f"The path '{dir_path_str}' is not a valid directory.")
            else:
                self.debug_window.log_message(f"No directory item found at row {row}, column 0.", "WARNING")
        except Exception as e:
            self.debug_window.log_message(f"Error opening directory for row {row}: {e}", "ERROR")
            QMessageBox.critical(self, "Error", f"Could not open directory. Error: {e}")

    def queue_selected_files_for_deletion(self):
        # Get the currently selected rows
        selected_items_indices = self.table.selectionModel().selectedRows() 
        self.debug_window.log_message(f"SELECTION DEBUG: Number of rows selected: {len(selected_items_indices)}", "INFO")
        
        if not selected_items_indices:
            QMessageBox.information(self, "No Selection", "Please select one or more files to queue for deletion.")
            return

        items_to_review = []
        self.debug_window.log_message(f"SELECTION DEBUG: Processing {len(selected_items_indices)} selected rows", "INFO")
        
        # Collect details about all selections for debugging
        selection_details = []
        for i, model_index in enumerate(selected_items_indices):
            row = model_index.row()
            selection_details.append(f"Selection {i+1}: Row {row}")
        self.debug_window.log_message(f"SELECTION DEBUG: Selection details: {', '.join(selection_details)}", "INFO")
        
        # For each selected row
        for i, model_index in enumerate(selected_items_indices):
            row = model_index.row()
            self.debug_window.log_message(f"SELECTION DEBUG: Processing selection {i+1}, row {row}", "INFO")
            
            # Skip if row is hidden
            if self.table.isRowHidden(row):
                self.debug_window.log_message(f"SELECTION DEBUG: Row {row} is hidden - skipping", "INFO")
                continue
            
            # Get all column items for this row to ensure we're looking at the right data
            try:
                dir_item = self.table.item(row, 0)  # Directory column
                file_item = self.table.item(row, 1)  # Filename column
                created_item = self.table.item(row, 2)  # Created column
                modified_item = self.table.item(row, 3)  # Modified column
                status_item = self.table.item(row, 4)  # Status column
                
                # Debug log all column values for this row
                dir_text = dir_item.text() if dir_item else "None"
                file_text = file_item.text() if file_item else "None"
                created_text = created_item.text() if created_item else "None"
                modified_text = modified_item.text() if modified_item else "None"
                status_text = status_item.text() if status_item else "None"
                
                self.debug_window.log_message(
                    f"SELECTION DEBUG: Row {row} data - Dir: {dir_text}, File: {file_text}, "
                    f"Created: {created_text}, Modified: {modified_text}, Status: {status_text}", "INFO")
                
                # Check if already queued
                if status_item and status_item.text() == "Queued for Deletion":
                    self.debug_window.log_message(f"SELECTION DEBUG: Row {row} already queued for deletion - skipping", "INFO")
                    continue
                
                # Verify both directory and filename exist
                if dir_item is not None and file_item is not None:
                    items_to_review.append((dir_text, file_text))
                    self.debug_window.log_message(f"SELECTION DEBUG: Row {row} - Added to review list: {dir_text} / {file_text}", "INFO")
                else:
                    self.debug_window.log_message(f"SELECTION DEBUG: Row {row} - Skipping: Missing dir or file item", "WARNING")
                    
            except Exception as e:
                self.debug_window.log_message(f"SELECTION DEBUG: Error processing row {row}: {str(e)}", "ERROR")

        self.debug_window.log_message(f"SELECTION DEBUG: Finished processing selections. Added {len(items_to_review)} items to review list", "INFO")
        
        if not items_to_review:
            QMessageBox.information(self, "No New Items", "No new files selected to queue (already queued or no selection).")
            return

        self.debug_window.log_message(f"SELECTION DEBUG: Creating ReviewDialog with {len(items_to_review)} items", "INFO")
        review_dialog = ReviewDialog(items_to_review, self)
        
        if review_dialog.exec(): 
            items_confirmed_for_queue = review_dialog.get_items_confirmed_for_queue()
            self.debug_window.log_message(f"SELECTION DEBUG: ReviewDialog returned {len(items_confirmed_for_queue)} confirmed items", "INFO")
            
            if items_confirmed_for_queue:
                for dir_path, file_name in items_confirmed_for_queue:
                    self.files_queued_for_deletion.add((dir_path, file_name))
                    # Update status in self.current_scanned_files_data
                    for data_entry in self.current_scanned_files_data:
                        if data_entry.get("directory") == dir_path and data_entry.get("filename") == file_name:
                            data_entry["status"] = "Queued for Deletion" 
                            break
                self.debug_window.log_message(f"Added {len(items_confirmed_for_queue)} items to deletion queue.", "INFO")
                self.populate_table_view() # Refresh table to show new statuses
                self.update_queue_counter() # Update counter after adding to queue
            else:
                self.debug_window.log_message("No items were ultimately confirmed for deletion queue.", "INFO")
        else: 
            self.debug_window.log_message("File queuing cancelled by user in review dialog.", "INFO")
        self.table.setCurrentItem(None)

    def clear_all_queued_deletions(self):
        if not self.files_queued_for_deletion:
            QMessageBox.information(self, "Nothing to Clear", "No files are currently queued for deletion.")
            return
        
        reply = QMessageBox.question(self, "Clear Queue",
                                     f"Are you sure you want to clear all {len(self.files_queued_for_deletion)} files from the deletion queue?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.files_queued_for_deletion.clear()
            # Reset status in self.current_scanned_files_data
            for data_entry in self.current_scanned_files_data:
                data_entry["status"] = "File Found" # Reset all, populate_table_view will handle it
            self.populate_table_view() # Refresh table
            self.update_queue_counter() # Update counter after clearing queue - resets to 0
            self.debug_window.log_message("Deletion queue cleared.", "INFO")
            QMessageBox.information(self, "Queue Cleared", "All files removed from the deletion queue.")

    def execute_queued_deletions_and_exit(self):
        if not self.files_queued_for_deletion:
            QMessageBox.information(self, "Nothing to Delete", "No files are queued for deletion. Closing application.")
            self.close()
            return

        reply = QMessageBox.warning(self, "Confirm Deletion",
                                     f"You are about to permanently send {len(self.files_queued_for_deletion)} files to the trash and then exit.\nThis action cannot be undone from within this application.\n\nAre you absolutely sure?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            self.debug_window.log_message("Execution of queued deletions cancelled by user.", "INFO")
            return

        QApplication.setOverrideCursor(Qt.WaitCursor)
        deleted_count = 0
        failed_count = 0
        
        # Create a list of (dir_path, file_name) from the set for iteration
        items_to_process_from_queue = list(self.files_queued_for_deletion)
        
        # For directory deletion logic, we need a list of Python files that will *remain*
        # i.e., all scanned files MINUS those in the current deletion queue.
        queued_keys = set(self.files_queued_for_deletion)
        py_files_not_in_queue = [
            item for item in self.current_scanned_files_data
            if (item.get("directory"), item.get("filename")) not in queued_keys
        ]

        for dir_path, file_name in items_to_process_from_queue:
            full_path = Path(dir_path) / file_name
            self.debug_window.log_message(f"Executing deletion for: {full_path}", "INFO")
            
            if self.scanner.delete_file_to_trash(str(full_path)):
                deleted_count += 1
                # Attempt to delete parent directory, passing the list of PY files *not* in the queue
                self.scanner.try_delete_empty_dir_to_trash(dir_path, py_files_not_in_queue)
            else:
                failed_count += 1
                self.debug_window.log_message(f"Failed to delete file: {full_path}", "ERROR")

        QApplication.restoreOverrideCursor()
        summary_message = f"Deletion process complete.\nSuccessfully sent to trash: {deleted_count} files.\nFailed to send to trash: {failed_count} files."
        self.debug_window.log_message(summary_message.replace("\n", " "), "INFO")
        QMessageBox.information(self, "Deletion Summary", summary_message)
        
        self.debug_window.log_message("Exiting application after executing deletions.", "INFO")
        
        # Set flag to indicate we're exiting after deletion
        self.is_exit_after_deletion = True
        
        # Clear the queue since deletions are processed
        self.files_queued_for_deletion.clear()
        
        self.close()

    def close_application_handler(self):
        """Handles the 'Close Application' button click and window close event."""
        # Skip confirmation if we've already confirmed deletion and are exiting
        if self.is_exit_after_deletion:
            return True
            
        if self.files_queued_for_deletion:
            reply = QMessageBox.question(self, "Unsaved Changes",
                                         "You have files queued for deletion. Are you sure you want to exit without executing them?",
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return False # Prevent closing
        self.close() # Proceed with closing
        return True

    def closeEvent(self, event):
        """Handle the main window close event (e.g., clicking X)."""
        self.debug_window.log_message("Application close event triggered.", "INFO")
        if not self.close_application_handler(): # Check if there are unsaved changes
            event.ignore() # Ignore the close event if user cancels
            return
            
        if self.debug_window and self.debug_window.isVisible():
            self.debug_window.close()
        
        logger.info("Application shutting down.")
        super().closeEvent(event)

    def update_queue_counter(self):
        """Update the queue counter display with the current number of queued files"""
        count = len(self.files_queued_for_deletion)
        self.queue_counter_label.setText(str(count))
        if count > 0:
            self.queue_counter_label.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
        else:
            self.queue_counter_label.setStyleSheet("color: gray; font-weight: normal; font-size: 14px;")

    def export_current_list(self):
        items_to_export = self.current_scanned_files_data
        if not items_to_export:
            QMessageBox.information(self, "No Items to Export", "No files are currently scanned.")
            return

        # Create file format selection dialog
        format_dialog = QDialog(self)
        format_dialog.setWindowTitle("Select Export Format")
        format_dialog.setMinimumWidth(300)
        format_layout = QVBoxLayout(format_dialog)
        
        format_label = QLabel("<b>Select Export Format:</b>")
        format_layout.addWidget(format_label)
        
        # Format selection buttons
        csv_button = QPushButton("CSV")
        xlsx_button = QPushButton("XLSX")
        txt_button = QPushButton("TXT")
        pdf_button = QPushButton("PDF")
        
        format_layout.addWidget(csv_button)
        format_layout.addWidget(xlsx_button)
        format_layout.addWidget(txt_button)
        format_layout.addWidget(pdf_button)
        
        selected_format = [None]  # Using list to store value by reference
        
        def set_format(fmt):
            selected_format[0] = fmt
            format_dialog.accept()
            
        csv_button.clicked.connect(lambda: set_format("csv"))
        xlsx_button.clicked.connect(lambda: set_format("xlsx"))
        txt_button.clicked.connect(lambda: set_format("txt"))
        pdf_button.clicked.connect(lambda: set_format("pdf"))
        
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(format_dialog.reject)
        format_layout.addWidget(cancel_button)
        
        format_dialog.exec()
        
        # If dialog was canceled or no format selected
        if not selected_format[0]:
            return
            
        file_format = selected_format[0]
        self.debug_window.log_message(f"Selected export format: {file_format}", "INFO")
        
        # Set default filename and filter based on selected format
        if file_format == "csv":
            default_filename = f"ScannedFiles_{QDate.currentDate().toString('yyyy-MM-dd')}.csv"
            file_filter = "CSV Files (*.csv);;All Files (*)"
        elif file_format == "xlsx":
            default_filename = f"ScannedFiles_{QDate.currentDate().toString('yyyy-MM-dd')}.xlsx"
            file_filter = "Excel Files (*.xlsx);;All Files (*)"
        elif file_format == "pdf":
            default_filename = f"ScannedFiles_{QDate.currentDate().toString('yyyy-MM-dd')}.pdf"
            file_filter = "PDF Files (*.pdf);;All Files (*)"
        else:  # txt format
            default_filename = f"ScannedFiles_{QDate.currentDate().toString('yyyy-MM-dd')}.txt"
            file_filter = "Text Files (*.txt);;All Files (*)"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Scanned Files List", 
            str(Path.home() / default_filename),
            file_filter
        )
        
        if not file_path:
            return
            
        try:
            if file_format == "csv":
                self.export_to_csv(file_path, items_to_export)
            elif file_format == "xlsx":
                self.export_to_xlsx(file_path, items_to_export)
            elif file_format == "pdf":
                self.export_to_pdf(file_path, items_to_export)
            else:  # txt format
                self.export_to_txt(file_path, items_to_export)
                
            QMessageBox.information(self, "Export Successful", f"Scanned files list exported to:\n{file_path}")
            self.debug_window.log_message(f"Scanned files list exported to {file_path}", "INFO")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Could not export scanned files list: {e}")
            self.debug_window.log_message(f"Error exporting scanned files list: {e}", "ERROR")
    
    def export_to_txt(self, file_path, items_to_export):
        """Export data to a text file."""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(f"Scanned Files List - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 40 + "\n")
            f.write("Files scanned:\n\n")
            for file_entry in items_to_export:
                f.write(f"Directory: {file_entry['directory']}\nFilename:  {file_entry['filename']}\n")
                f.write(f"Created: {file_entry['created']}\nModified: {file_entry['modified']}\n")
                f.write(f"Status: {file_entry['status']}\n---\n")
    
    def export_to_csv(self, file_path, items_to_export):
        """Export data to a CSV file."""
        import csv
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Directory', 'Filename', 'Created', 'Modified', 'Status'])
            for file_entry in items_to_export:
                writer.writerow([
                    file_entry['directory'],
                    file_entry['filename'],
                    file_entry['created'],
                    file_entry['modified'],
                    file_entry['status']
                ])
    
    def export_to_xlsx(self, file_path, items_to_export):
        """Export data to an Excel file."""
        try:
            import openpyxl
            from openpyxl.styles import Font
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Scanned Files"
            
            # Add header row with bold font
            headers = ['Directory', 'Filename', 'Created', 'Modified', 'Status']
            bold_font = Font(bold=True)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = bold_font
            
            # Add data rows
            for row_num, file_entry in enumerate(items_to_export, 2):
                ws.cell(row=row_num, column=1, value=file_entry['directory'])
                ws.cell(row=row_num, column=2, value=file_entry['filename'])
                ws.cell(row=row_num, column=3, value=file_entry['created'])
                ws.cell(row=row_num, column=4, value=file_entry['modified'])
                ws.cell(row=row_num, column=5, value=file_entry['status'])
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
        except ImportError:
            raise Exception("Excel export requires the openpyxl module. Please install it with 'pip install openpyxl'.")
    
    def export_to_pdf(self, file_path, items_to_export):
        """Export data to a PDF file."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            
            doc = SimpleDocTemplate(file_path, pagesize=landscape(letter))
            
            # Create the data for the table
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            data = [['Directory', 'Filename', 'Created', 'Modified', 'Status']]
            
            for file_entry in items_to_export:
                data.append([
                    file_entry['directory'],
                    file_entry['filename'],
                    file_entry['created'],
                    file_entry['modified'],
                    file_entry['status']
                ])
            
            # Create the table
            table = Table(data, repeatRows=1)
            
            # Style the table
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            # Create title
            title = Paragraph(f"Scanned Files List - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", title_style)
            
            # Build the PDF
            elements = [title, table]
            doc.build(elements)
        except ImportError:
            raise Exception("PDF export requires the reportlab module. Please install it with 'pip install reportlab'.")


def main():
    # Remove deprecated high DPI scaling attributes
    # Modern Qt versions handle high DPI scaling automatically
    app = QApplication(sys.argv)
    main_window = FileScannerApp()
    main_window.show()
    exit_code = app.exec()
    sys.exit(exit_code)


if __name__ == "__main__":
    if log_file_path.parent: 
        log_file_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info("================ Application Starting ================")
    main()
    logger.info("================ Application Exited =================")

